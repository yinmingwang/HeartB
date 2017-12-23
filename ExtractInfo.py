import openpyxl
import xlrd
import re
from Patient import Patient

def ExtractInfo(inputstr):
    inputstr = inputstr.strip()
    KeyVarible = ['AO根部', 'LA', 'RV常规', 'IVS', 'LVDd', 'LVDs', 'LVPW', 'RA横径', 'LVEF', 'AV', 'MVE', 'MVA']
    #print(inputstr)
    VaribleDic = {}
    for item in KeyVarible:
        content = re.findall(u'%s([\W*\s*\d*\w*]*?)[，：。]' % (item), inputstr)
        VaribleDic[item] = content
        #print(item,content)
    return VaribleDic
def  Left_Ventricle(inputstr):
    inputstr = inputstr.strip()
    str_list = ['减退','未见?异常']
    Systolic_flag = None
    Diastolic_flag = None
    Systolic_content = re.findall(u'(左心?室收缩功能[\W*\s*\d*\w*]*?[，：。、；])', inputstr)
    Diastolic_content = re.findall(u'左心?室[\W*\s*\d*\w*]*?(舒张功能[\W*\s*\d*\w*]*?[，：。、；])', inputstr)
    if len(Systolic_content) > 0:
        #print(Systolic_content)
        for index,item in enumerate(str_list):
            sub_content = re.findall(u'%s' % (item), Systolic_content[0])
            if len(sub_content) > 0:
                Systolic_flag = index
                break
    if len(Diastolic_content) > 0:
        #print(Diastolic_content)
        for index_one, item_one in enumerate(str_list):
            sub_content_one = re.findall(u'%s' % (item_one), Diastolic_content[0])
            if len(sub_content_one) > 0:
                Diastolic_flag = index_one
                break
    return Systolic_flag, Diastolic_flag

def AortaRoot(inputstr):
    inputstr = inputstr.strip()
    AortaRoot_content = re.findall(u'主动脉(?:根部)?[^瓣][\W*\s*\d*\w*]*?(?:[^壁瓣腔]搏动[\W*\s*\d*\w*]*?)?[，：。、；]',inputstr)
    root_des = ['不宽','([增稍]?增?宽)|不好']
    pulse_des = ['尚?[好可]','([稍减]?低平?)|呈弓背样']
    root_flag = None
    pulse_flag = None
    abnormal = None
    if len(AortaRoot_content) > 0:
        for index_r,item_r in enumerate(root_des):
            sub_root_des = re.findall(u'主动脉(?:根部)?[^瓣][\W*\s*\d*\w*]*?%s' % (item_r), AortaRoot_content[0])
            if len(sub_root_des) > 0:
                root_flag = index_r
                break
        for index_p, item_p in enumerate(pulse_des):
            sub_pulse_des = re.findall(u'搏动[\W*\s*\d*\w*]*?%s' % (item_p), AortaRoot_content[0])
            if len(sub_pulse_des) > 0:
                pulse_flag = index_p
                break
        if root_flag == 0 and pulse_flag == 0:
            abnormal = 0
        if root_flag == 1 or pulse_flag == 1:
            abnormal = 1
        print(AortaRoot_content,abnormal,root_flag,  pulse_flag)
    else:
        print('未找到相关字段')
    return root_flag, pulse_flag, abnormal

def  Aortic_Vaives(inputstr):
    inputstr = inputstr.strip()
    Aortic_Vaives_content = re.findall(u'主动脉瓣[\W*\s*\d*\w*]*?开放关闭[\W*\s*\d*\w*]*?',inputstr)
    if len(Aortic_Vaives_content) > 0:
        print(Aortic_Vaives_content)
    else:
        print('未找到相关字段')
def Left_Heart_System(inputstr):
    inputstr = inputstr.strip()
    Left_Heart_house_des=['(?:未见增大|不大)','稍?增?[大高]']
    Left_Heart_ventricle_des=['未见增大','稍?增?大']
    Left_Combustor_Wall_des=['未见增厚','变[薄厚]']
    Left_Combustor_Wall_pulse_des=['尚?[好可]','(?:明显|稍|普遍)?减弱']
    Left_Heart_house_flag = None
    Left_Heart_ventricle_flag = None
    Left_Combustor_Wall_incrassation_flag = None
    Left_Combustor_Wall_pulse_flag = None
    all_Left_Combustor_Wall = None
    abnormal = None
    Left_Heart_System_content = re.findall(u'(?:左房|左室)[^收缩功能][\W\s\d\w]*室壁[\W\s\d\w]*?[，：。、；]',inputstr)
    if len(Left_Heart_System_content) > 0:
        print( Left_Heart_System_content)
        for index_h,item_h in enumerate( Left_Heart_house_des):
            Left_Heart_house_content = re.findall(u'左房[\W\s\d\w]*?腔?[\W\s\d\w]*?%s' % (item_h),Left_Heart_System_content[0])
            if len(Left_Heart_house_content) > 0:
                Left_Heart_house_flag = index_h
                print(Left_Heart_house_content, Left_Heart_house_flag)
                break
        for index_v,item_v in enumerate(Left_Heart_ventricle_des):
            Left_Heart_ventricle_content =  re.findall(u'左室[\W\s\d\w]*[^房]腔?%s' % (item_v),Left_Heart_System_content[0])
            if len(Left_Heart_ventricle_content) > 0:
                Left_Heart_ventricle_flag = index_v
                break
        for index_w, item_w in enumerate(Left_Combustor_Wall_des):
            Left_Combustor_Wall_content = re.findall(u'[左余]?[余左]?[^右]室[\W\s\d\w]*?壁[\W\s\d\w]*?%s' % (item_w),Left_Heart_System_content[0])
            if len(Left_Combustor_Wall_content) > 1:
                for sub_index, sub_item in enumerate(Left_Combustor_Wall_des):
                    sub_Left_Combustor_Wall_content = re.findall(u'[左余]?[余左]?[^右]室[\W\s\d\w]*?壁[\W*\s\d\w]*?%s' % (sub_index),
                                                                 Left_Combustor_Wall_content[1])
                    if len(sub_Left_Combustor_Wall_content) > 0:
                        Left_Combustor_Wall_incrassation_flag = sub_index
                        break
            if len(Left_Combustor_Wall_content) > 0 and len(Left_Combustor_Wall_content) < 2:
                Left_Combustor_Wall_incrassation_flag = index_w
                break
        for index_p, item_p in enumerate(Left_Combustor_Wall_pulse_des):
            Left_Combustor_Wall_pulse_content = re.findall(u'[左余]?[^右]室[\W\s\d\w]*?壁[\W\s\d\w]*?[运搏]动%s' % (item_p), Left_Heart_System_content[0])
            if len(Left_Combustor_Wall_pulse_content) > 0:
                Left_Combustor_Wall_pulse_flag = index_p
                break
    else:
        print('未找到相关字段')
    if Left_Combustor_Wall_pulse_flag == 1 or  Left_Combustor_Wall_incrassation_flag == 1:
        all_Left_Combustor_Wall = 1
    if Left_Combustor_Wall_pulse_flag == 0 and Left_Combustor_Wall_incrassation_flag == 0:
        all_Left_Combustor_Wall = 0
    if Left_Heart_house_flag == 0 and Left_Heart_ventricle_flag == 0 and all_Left_Combustor_Wall == 0:
        abnormal = 0
    if Left_Heart_house_flag == 1 or Left_Heart_ventricle_flag == 1 or all_Left_Combustor_Wall == 1:
        abnormal = 1
    print(abnormal, Left_Heart_house_flag, Left_Heart_ventricle_flag,all_Left_Combustor_Wall)
    return abnormal, Left_Heart_house_flag,Left_Heart_ventricle_flag,all_Left_Combustor_Wall

def Right_Heart_System(inputstr):
    inputstr = inputstr.strip()
    Right_Heart_house_des = ['(?:未见增大|不大)', '稍?增?[大高]']
    Right_Heart_ventricle_des = ['未见增大', '稍?增?大']
    Right_Heart_house_flag = None
    Right_Heart_ventricle_flag = None
    for index_h, item_h in enumerate(Right_Heart_house_des):
        Right_Heart_house_content = re.findall(u'右房[\W\s\d\w]*?腔?[\W\s\d\w]*?%s' % (item_h),
                                               inputstr)
        if len(Right_Heart_house_des) > 0:
            Right_Heart_house_flag = index_h
            break
    for index_v,item_v in enumerate(Right_Heart_ventricle_des):
        Right_Heart_ventricle_content =  re.findall(u'右室[\W\s\d\w]*[^房]腔?%s' % (item_v), inputstr)
        if len(Right_Heart_ventricle_content) > 0:
            Right_Heart_ventricle_flag = index_v
    return Right_Heart_house_flag, Right_Heart_ventricle_flag

if __name__ == '__main__':
    wb = xlrd.open_workbook('心脏B超.xls')
    ws = wb.sheet_by_name('Sheet1')
    rows = ws.nrows
    cols = ws.ncols
    Patients = []
    Patients_item = ['AO', 'LA', 'RV', 'IVS', 'LVDd', 'LVDs','LVPW', 'RA', 'LVEF', 'AV', 'MVE', 'MVA']
    for row in range(1,rows):
        PatientName = ws.cell(row, 0).value
        print(PatientName)
        ReportDescribe = ws.cell(row, 1).value
        ReportDiagnose = ws.cell(row, 2).value
        #MainReportDescribe = ReportDescribe+"\n"+ReportDiagnose
        #print(MainReportDescribe)
        Var = ExtractInfo(ReportDescribe)
        Left_Systolic_flag, Left_Diastolic_flag = Left_Ventricle(ReportDiagnose)
        #root_flag, pulse_flag, abnormal = AortaRoot(ReportDescribe)
        Left_Heart_System(ReportDescribe)
        #Aortic_Vaives(ReportDescribe)
        #Var = ExtractInfo(MainReportDescribe)
        #Left_Systolic_flag, Left_Diastolic_flag = Left_Ventricle(MainReportDescribe)
        #root_flag, pulse_flag, abnormal = AortaRoot(MainReportDescribe)
        #print( Left_Systolic_flag,  Left_Diastolic_flag)
        patient = Patient(PatientNmae=PatientName)
        if len(Var['AO根部']) > 0:
            patient.AO = Var['AO根部'][0]
        if len(Var['LA']) > 0:
            patient.LA = Var['LA'][0]
        if len(Var['RV常规']) > 0:
            patient.RV = Var['RV常规'][0]
        if len(Var['IVS']) > 0:
            patient.IVS = Var['IVS'][0]
        if len(Var['LVDd']) > 0:
            patient.LVDd = Var['LVDd'][0]
        if len(Var['LVDs']) > 0:
            patient.LVDs = Var['LVDs'][0]
        if len(Var['LVPW']) > 0:
            patient.LVPW = Var['LVPW'][0]
        if len(Var['RA横径']) > 0:
            patient.RA = Var['RA横径'][0]
        if len(Var['LVEF']) > 0:
            patient.LVEF = Var['LVEF'][0]
        if len(Var['AV']) > 0:
            patient.AV = Var['AV'][0]
        if len(Var['MVE']) > 0:
            patient.MVE = Var['MVE'][0]
        if len(Var['MVA']) > 0:
            patient.MVA = MVA = Var['MVA'][0]
        patient.Left_ventricle_systolic = Left_Systolic_flag
        patient.Left_ventricle_diastolic = Left_Diastolic_flag
        #patient.print()
        Patients.append(patient)







